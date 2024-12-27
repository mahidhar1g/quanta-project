import pandas as pd
import os
from openpyxl.styles import Font, Alignment, PatternFill

def export_to_excel(data, output_file="breakout_analysis.xlsx"):
    """
    Export the DataFrame to a styled Excel file with additional details.

    Args:
        data (pd.DataFrame): DataFrame containing the breakout analysis results.
        output_file (str): Name of the output Excel file.

    Returns:
        str: Path to the generated Excel file.
    """
    # Define the exports directory
    export_dir = os.path.join(os.getcwd(), "exports")
    if not os.path.exists(export_dir):
        os.makedirs(export_dir)

    # Full file path
    file_path = os.path.join(export_dir, output_file)

    # Create Excel writer and write data
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        # Write the main data
        data.to_excel(writer, index=False, sheet_name="Breakout Analysis")
        
        # Optionally, you can add introductory or metadata sheets
        # For example, a description sheet
        intro_data = {
            "Column Name": [
                "date", "volume", "close", "price_change_pct_to_prev_day",
                "closing_date_after_holding", "close_after_holding", "holding_return"
            ],
            "Description": [
                "Date of the breakout event",
                "Trading volume on the breakout day",
                "Closing price on the breakout day",
                "Percentage change in price compared to the previous day",
                "Date after the specified holding period",
                "Closing price after the holding period",
                "Percentage return after the holding period"
            ]
        }
        pd.DataFrame(intro_data).to_excel(writer, index=False, sheet_name="Details")

        # Styling headers (openpyxl allows this after writing the data)
        workbook = writer.book
        sheet = writer.sheets["Breakout Analysis"]

        # Apply header styles
        for col in sheet.iter_cols(min_row=1, max_row=1, min_col=1, max_col=len(data.columns)):
            for cell in col:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Return the full file path for the route
    return file_path
